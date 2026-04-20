import streamlit as st
import pandas as pd
import numpy as np
import os, io, re
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Diseño Hidráulico Pro",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ══════════════════════════════════════════════════════════════════════════════
# OCULTAR DEPLOY Y MENÚ
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
#MainMenu {visibility: hidden !important;}
header[data-testid="stHeader"] {visibility: hidden !important;}
[data-testid="stToolbar"] {display: none !important;}
.stDeployButton {display: none !important;}
footer {visibility: hidden !important;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# ESTILOS GLOBALES
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@500;600;700&family=Exo+2:ital,wght@0,300;0,400;0,500;1,300&family=Share+Tech+Mono&display=swap" rel="stylesheet">', unsafe_allow_html=True)

st.markdown("""
<style>
div[data-testid="stTabs"] > div > div[role="tablist"] {
  background: #f0f4fa !important;
  border-bottom: 2px solid #c5d5ea !important;
  padding: 0 20px !important;
}
div[data-testid="stTabs"] button[role="tab"] {
  color: #5a7aaa !important;
  -webkit-text-fill-color: #5a7aaa !important;
  font-family: 'Rajdhani', sans-serif !important;
  font-size: 0.82rem !important; font-weight: 600 !important;
  letter-spacing: 2px !important; text-transform: uppercase !important;
  border: none !important; border-bottom: 3px solid transparent !important;
  padding: 13px 26px !important; background: transparent !important;
  border-radius: 0 !important;
}
div[data-testid="stTabs"] button[role="tab"]:hover {
  color: #1a56b0 !important; -webkit-text-fill-color: #1a56b0 !important;
  border-bottom-color: rgba(26,86,176,0.3) !important;
}
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
  color: #1a56b0 !important; -webkit-text-fill-color: #1a56b0 !important;
  border-bottom-color: #1a56b0 !important;
  background: rgba(26,86,176,0.07) !important;
}
div[data-testid="stTabs"] button[role="tab"] p,
div[data-testid="stTabs"] button[role="tab"] span,
div[data-testid="stTabs"] button[role="tab"] div {
  color: inherit !important; -webkit-text-fill-color: inherit !important;
  font-size: inherit !important; font-family: inherit !important;
  font-weight: inherit !important; letter-spacing: inherit !important;
  text-transform: inherit !important;
}
.dh-card {
  background: #ffffff; border: 1px solid #c5d5ea;
  border-top: 3px solid #1a56b0; border-radius: 4px;
  padding: 22px 26px; margin-bottom: 18px; position: relative;
}
.dh-card::after {
  content: ''; position: absolute; top: 0; right: 0;
  width: 36px; height: 36px;
  border-top: 3px solid #2a6fdb; border-right: 3px solid #2a6fdb;
  border-radius: 0 4px 0 0;
}
.dh-card-title {
  font-family: 'Rajdhani', sans-serif; font-size: 0.9rem; font-weight: 700;
  color: #1a56b0; letter-spacing: 3px; text-transform: uppercase;
  margin-bottom: 16px; padding-bottom: 9px; border-bottom: 1px solid #c5d5ea;
}
.dh-metric {
  background: #eef4fb; border: 1px solid #c5d5ea;
  border-left: 3px solid #1a56b0; padding: 11px 15px; border-radius: 2px;
  font-family: 'Share Tech Mono', monospace;
}
.dh-metric-label { font-size: 0.63rem; color: #5a7aaa; letter-spacing: 2px; text-transform: uppercase; }
.dh-metric-value { font-size: 1.25rem; color: #1a56b0; margin-top: 2px; }
.dh-metric-unit  { font-size: 0.63rem; color: #5a7aaa; }
.dh-sep { border: none; height: 1px;
  background: linear-gradient(90deg, transparent, #2a6fdb, transparent);
  margin: 18px 0; }
[data-testid="stSlider"] > div > div > div { background: #1a56b0 !important; }
[data-testid="stSelectbox"] > div > div,
[data-testid="stRadio"] > div {
  background: #f5f8fd !important; border-color: #c5d5ea !important;
  color: #1a2a4a !important; font-family: 'Exo 2', sans-serif !important;
}
[data-testid="stRadio"] label { color: #1a2a4a !important; }
[data-testid="stButton"] > button {
  background: transparent !important; border: 1px solid #1a56b0 !important;
  color: #1a56b0 !important; font-family: 'Rajdhani', sans-serif !important;
  font-weight: 600 !important; letter-spacing: 2px !important;
  text-transform: uppercase !important; font-size: 0.8rem !important;
  padding: 10px 22px !important; border-radius: 2px !important;
  transition: all 0.2s !important;
}
[data-testid="stButton"] > button:hover {
  background: rgba(26,86,176,0.08) !important;
  box-shadow: 0 0 18px rgba(26,86,176,0.2) !important;
}
[data-testid="stButton"] > button[kind="primary"] {
  background: #1a56b0 !important; color: #ffffff !important;
}
[data-testid="stDownloadButton"] > button {
  background: #f0f4fa !important; border: 1px solid #c5d5ea !important;
  color: #1a2a4a !important; font-family: 'Rajdhani', sans-serif !important;
  font-weight: 600 !important; letter-spacing: 1.5px !important;
  text-transform: uppercase !important; font-size: 0.76rem !important;
  border-radius: 2px !important; transition: all 0.2s !important;
}
[data-testid="stDownloadButton"] > button:hover {
  border-color: #1a56b0 !important; color: #1a56b0 !important;
}
.res-table { width: 100%; border-collapse: collapse; font-family: 'Exo 2', sans-serif; font-size: 0.82rem; }
.res-table th {
  background: #1e3a6e; color: #ffffff; padding: 10px 14px; text-align: center;
  font-family: 'Rajdhani', sans-serif; letter-spacing: 1.5px; font-size: 0.76rem;
  border: 1px solid #2a4a7a; text-transform: uppercase;
}
.res-table td { padding: 9px 12px; border: 1px solid #c5d5ea; text-align: center; color: #1a2a4a; }
.res-table tr:nth-child(even) td { background: #f0f5fb; }
.res-table tr:hover td { background: #ddeaf8; }
.res-table td.left-align { text-align: left; font-weight: 600; color: #1a3a6a; }
.res-table td.highlight-hw { color: #1a56b0; font-weight: 700; }
.res-table td.highlight-dw { color: #1a7a9a; font-weight: 700; }
label, [data-testid="stWidgetLabel"] p {
  font-family: 'Exo 2', sans-serif !important; color: #5a7aaa !important;
  font-size: 0.78rem !important; letter-spacing: 1px !important;
}
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f0f4fa; }
::-webkit-scrollbar-thumb { background: #c5d5ea; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #1a56b0; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style="background:linear-gradient(135deg,#1e3a6e 0%,#2a5aaa 50%,#1a3060 100%);
  border-bottom:2px solid #1a56b0;padding:22px 36px 18px;border-radius:6px;margin-bottom:0;">
  <span style="font-family:'Rajdhani',sans-serif;font-size:11px;color:#6a9fdf;
    letter-spacing:0.22em;text-transform:uppercase;margin-bottom:8px;display:block;">
    Análisis de Pérdidas de Carga</span>
  <span style="font-family:'Rajdhani',sans-serif;font-size:2.4rem;font-weight:700;
    color:#ffffff;margin:0 0 6px;line-height:1.15;display:block;">Diseño Hidráulico</span>
  <span style="font-family:'Exo 2',sans-serif;font-size:0.87rem;font-style:italic;
    font-weight:300;color:#a8c8f0;margin:0 0 14px;display:block;">
    Modelamiento computacional de flujo interno en tuberías a presión</span>
  <span style="font-family:'Share Tech Mono',monospace;font-size:0.7rem;color:#7aaad0;
    padding-top:11px;border-top:1px solid rgba(255,255,255,0.15);line-height:1.8;display:block;">
    Hecho por: <strong style="color:#ffffff;">Nombre Completo</strong>
    &nbsp;·&nbsp; Universidad Nacional del Altiplano · Puno 2026</span>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data
def cargar_datos():
    df_agua = pd.read_excel("datos.xlsx", sheet_name="PropiedadesAgua", decimal=',')
    df_rug  = pd.read_excel("datos.xlsx", sheet_name="Rugosidad",        decimal='.')
    df_tubo = pd.read_excel("datos.xlsx", sheet_name="TuberiaData",      decimal='.')
    for col in ['Temp','Visc_Cinem','Densidad','Visc_Dinam']:
        df_agua[col] = pd.to_numeric(df_agua[col], errors='coerce')
    df_rug['Rugosidad']   = pd.to_numeric(df_rug['Rugosidad'],   errors='coerce')
    df_tubo['Diametro_m'] = pd.to_numeric(df_tubo['Diametro_m'], errors='coerce')
    df_tubo['Area_m2']    = pd.to_numeric(df_tubo['Area_m2'],    errors='coerce')
    return df_agua, df_rug, df_tubo

df_agua, df_rug, df_tubo = cargar_datos()

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
if 'lista_resultados' not in st.session_state:
    st.session_state.lista_resultados = []
if 'go_to_tab' not in st.session_state:
    st.session_state.go_to_tab = None

def jump_to_tab(tab_index: int):
    js = f"""<script>(function(){{var a=0;function t(){{a++;var tabs=window.parent.document.querySelectorAll('[data-testid="stTabs"] [role="tab"]');if(tabs.length>{tab_index}){{tabs[{tab_index}].click();}}else if(a<30){{setTimeout(t,100);}}}}setTimeout(t,80);}})();</script>"""
    st.components.v1.html(js, height=0)

if st.session_state.go_to_tab == 1:
    jump_to_tab(1)
    st.session_state.go_to_tab = None

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def safe_sheet_name(name: str, max_len: int = 31) -> str:
    return re.sub(r'[\\/*?:\[\]/]', '-', name)[:max_len]

def detectar_diam_nom(nombre_tubo: str) -> str:
    n = nombre_tubo.upper()
    if '3/4' in n:
        return '3/4"'
    if '1"' in n or n.strip().endswith(' 1'):
        return '1"'
    return '1/2"'

def detectar_material_tubo(nombre_tubo: str, df_rug: pd.DataFrame) -> str:
    """Detecta material comparando sin caracteres especiales para soportar F°G°."""
    n_clean = re.sub(r'[°\s]', '', nombre_tubo.upper())
    for mat in df_rug['Material']:
        mat_clean = re.sub(r'[°\s]', '', mat.upper())
        if mat_clean and mat_clean in n_clean:
            return mat
    if 'PVC' in nombre_tubo.upper():
        return 'PVC'
    if 'HDPE' in nombre_tubo.upper():
        return 'HDPE'
    # FG / FG / Fierro galvanizado
    if 'FG' in n_clean or 'FIERRO' in nombre_tubo.upper() or 'GALV' in nombre_tubo.upper():
        # buscar el material FG en df_rug
        for mat in df_rug['Material']:
            if 'FG' in re.sub(r'[°\s]', '', mat.upper()):
                return mat
    return df_rug['Material'].iloc[0]

def calcular(nombre_tubo, material, rugosidad, diametro, area, caudal_ls, longitud, temp, visc_cin):
    Q_m3s = caudal_ls / 1000
    V     = Q_m3s / area
    Re    = (diametro * V) / visc_cin
    eD    = rugosidad / diametro
    f     = 0.25 / (np.log10((rugosidad/(3.71*diametro)) + (5.74/(Re**0.9))))**2
    lnRe  = np.log(Re); lneD = np.log(eD)
    C_H   = (197.17 - 25.79*lnRe - 5.41*lneD + 0.4464*lnRe**2
             - 3.39*lneD**2 - 5.086*lnRe*lneD + 0.041*lnRe**3
             + 0.124*lneD**3 + 0.39*lnRe*lneD**2 + 0.3757*lnRe**2*lneD)
    K_HW  = (10.67 * longitud) / ((C_H**1.852) * (diametro**4.87))
    Hf_HW = K_HW * (Q_m3s**1.852)
    K_DW  = (0.08263 * f * longitud) / (diametro**5)
    Hf_DW = K_DW * (Q_m3s**2)
    return {
        "nombre_tubo": nombre_tubo, "material": material,
        "diam_nom": detectar_diam_nom(nombre_tubo), "caudal_ls": caudal_ls,
        "diametro_m": diametro, "area_m2": area, "longitud": longitud,
        "temp": temp, "visc_cin": visc_cin, "rugosidad": rugosidad,
        "Velocidad": V, "Re": Re, "rel_eD": eD, "f": f, "C_H": C_H,
        "K_HW": K_HW, "Hf_HW": Hf_HW, "K_DW": K_DW, "Hf_DW": Hf_DW,
    }

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — recibe datos_editados para incluirlos en la exportación
# ══════════════════════════════════════════════════════════════════════════════
def _border(color="8EAADB"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):   return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="1F2937", size=9, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

XL = dict(
    title_bg="12294A", title_fg="FFFFFF",
    head1_bg="1A4A7A", head1_fg="FFFFFF",
    head2_bg="2563A8", head2_fg="FFFFFF",
    pvc_bg="D6E4F5",   pvc_fg="0D2137",
    fg_bg ="F5E6D0",   fg_fg ="5A2D00",
    hdpe_bg="D5EAD8",  hdpe_fg="1A3D1F",
    row_even="EDF3FA", row_odd="FFFFFF",
    border_c="8EAADB",
)

def build_detalle_sheet(ws, r, idx_num):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "ANÁLISIS DE PÉRDIDA DE CARGA — DISEÑO HIDRÁULICO"
    c.font = _font(bold=True, size=13, color=XL["title_fg"])
    c.fill = _fill(XL["title_bg"]); c.alignment = _align(); c.border = _border(XL["border_c"])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Reporte Detallado  |  Tubería: {r['nombre_tubo']}  ·  Material: {r['material']}  ·  Ø nominal: {r['diam_nom']}"
    ws["A2"].font = _font(size=9, color="3A5A8A"); ws["A2"].fill = _fill("D6E4F5")
    ws["A2"].alignment = _align(); ws["A2"].border = _border(XL["border_c"])
    ws.row_dimensions[2].height = 16

    for i, (h, w) in enumerate(zip(["Parámetro","Valor","Unidad","Observación"],[32,18,14,30]), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = _font(bold=True, color=XL["head2_fg"]); cell.fill = _fill(XL["head2_bg"])
        cell.alignment = _align(); cell.border = _border(XL["border_c"])
    ws.row_dimensions[3].height = 18

    rows_data = [
        ("Caudal",                f"{r['caudal_ls']} L/s = {r['caudal_ls']/1000:.4f} m³/s",   "L/s / m³/s", ""),
        ("Diámetro interno",      f"{r['diametro_m']*1000:.1f} mm = {r['diametro_m']:.4f} m",  "mm / m",     ""),
        ("Área transversal",      f"{r['area_m2']:.5f}",                                         "m²",         ""),
        ("Longitud del tramo",    f"{r['longitud']:.2f}",                                         "m",          ""),
        ("Temperatura del agua",  f"{r['temp']}",                                                 "°C",         ""),
        ("Viscosidad cinemática", f"{r['visc_cin']:.8f}",                                         "m²/s",       "Interpolado de tablas"),
        ("Velocidad media (v)",   f"{r['Velocidad']:.4f}",                                        "m/s",        "v = Q/A"),
        ("Número de Reynolds",    f"{r['Re']:.0f}",                                               "—",          "Re = v·D/ν"),
        ("Rugosidad absoluta ε",  f"{r['rugosidad']:.7f}",                                        "m",          ""),
        ("Relación ε/D",          f"{r['rel_eD']:.7f}",                                           "—",          ""),
        ("Factor de fricción f",  f"{r['f']:.5f}",                                                "—",          "Swamee-Jain"),
        ("C Hazen-Williams",      f"{r['C_H']:.3f}",                                              "—",          "Ec. regresión ln"),
        ("K Hazen-Williams",      f"{r['K_HW']:.4f}",                                             "s^1.852/m2", ""),
        ("Hf Hazen-Williams",     f"{r['Hf_HW']:.5f}",                                            "m",          "Hf = K*Q^1.852"),
        ("K Darcy-Weisbach",      f"{r['K_DW']:.4f}",                                             "s2/m5",      ""),
        ("Hf Darcy-Weisbach",     f"{r['Hf_DW']:.5f}",                                            "m",          "Hf = K·Q²"),
    ]
    for ri, (param, val, unit, obs) in enumerate(rows_data, start=4):
        bg = XL["row_even"] if ri % 2 == 0 else XL["row_odd"]
        for ci, v in enumerate([param, val, unit, obs], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = _fill(bg); cell.border = _border(XL["border_c"])
            cell.font = _font(color="1F2937", bold=(ci==1))
            cell.alignment = _align(h="left" if ci in (1,4) else "center")
        ws.row_dimensions[ri].height = 16

    cr = len(rows_data) + 5
    ws.merge_cells(f"A{cr}:D{cr}")
    c = ws.cell(row=cr, column=1, value="COMPARACIÓN DE MÉTODOS")
    c.font = _font(bold=True, size=10, color=XL["title_fg"])
    c.fill = _fill(XL["head1_bg"]); c.alignment = _align(); c.border = _border(XL["border_c"])
    ws.row_dimensions[cr].height = 18
    for ci, h in enumerate(["Método","Caudal (m³/s)","Velocidad (m/s)","Coeficiente","K","Hf (m)"], 1):
        cell = ws.cell(row=cr+1, column=ci, value=h)
        cell.font = _font(bold=True, color="FFFFFF"); cell.fill = _fill(XL["head2_bg"])
        cell.alignment = _align(); cell.border = _border(XL["border_c"])
    ws.row_dimensions[cr+1].height = 16
    mat_map = {"PVC":XL["pvc_bg"],"FG":XL["fg_bg"],"HDPE":XL["hdpe_bg"]}
    rbg = mat_map.get(r.get("material",""), XL["row_even"])
    for di, row in enumerate([
        ("Hazen-Williams", r['caudal_ls']/1000, r['Velocidad'], r['C_H'], r['K_HW'], r['Hf_HW']),
        ("Darcy-Weisbach", r['caudal_ls']/1000, r['Velocidad'], r['f'],   r['K_DW'], r['Hf_DW']),
    ], start=cr+2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=di, column=ci,
                           value=v if ci==1 else f"{v:.5f}" if isinstance(v,float) else v)
            cell.fill = _fill(rbg); cell.border = _border(XL["border_c"])
            cell.font = _font(color="1F2937", bold=(ci==1))
            cell.alignment = _align(h="left" if ci==1 else "center")
        ws.row_dimensions[di].height = 16


def build_resumen_sheet(ws, lista, datos_editados):
    ws.sheet_view.showGridLines = False
    caudales   = [0.5, 0.9, 1.5]
    materiales = ["PVC", "FG", "HDPE"]
    diams_nom  = ["1/2\"", "3/4\"", "1\""]
    idx        = {(r["caudal_ls"], r["material"], r["diam_nom"]): r for r in lista}
    mat_fill   = {"PVC":XL["pvc_bg"],"FG":XL["fg_bg"],"HDPE":XL["hdpe_bg"]}
    mat_font   = {"PVC":XL["pvc_fg"],"FG":XL["fg_fg"],"HDPE":XL["hdpe_fg"]}

    total_cols = 1 + 27
    last_col   = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "ANÁLISIS DE PÉRDIDA DE CARGA — DISEÑO HIDRÁULICO · CUADRO CONSOLIDADO"
    c.font = _font(bold=True, size=13, color=XL["title_fg"])
    c.fill = _fill(XL["title_bg"]); c.alignment = _align(); c.border = _border()
    ws.row_dimensions[1].height = 26

    # Fila 2: caudales
    ws.cell(row=2,column=1,value="CAUDAL (L/S)").font = _font(bold=True,color="1F2937")
    ws.cell(row=2,column=1).fill = _fill("D6E4F5")
    ws.cell(row=2,column=1).alignment = _align(); ws.cell(row=2,column=1).border = _border()
    col = 2
    for q in caudales:
        ws.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+8)
        c = ws.cell(row=2,column=col,value=q)
        c.font = _font(bold=True,size=11,color=XL["head1_fg"])
        c.fill = _fill(XL["head1_bg"]); c.alignment = _align()
        for cc in range(col,col+9): ws.cell(row=2,column=cc).border = _border()
        col += 9
    ws.row_dimensions[2].height = 20

    # Fila 3: materiales
    ws.cell(row=3,column=1,value="MATERIAL").font = _font(bold=True,color="1F2937")
    ws.cell(row=3,column=1).fill = _fill("D6E4F5")
    ws.cell(row=3,column=1).alignment = _align(); ws.cell(row=3,column=1).border = _border()
    col = 2
    for _ in caudales:
        for mat in materiales:
            ws.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+2)
            c = ws.cell(row=3,column=col,value=mat)
            c.font = _font(bold=True,color=mat_font[mat])
            c.fill = _fill(mat_fill[mat]); c.alignment = _align()
            for cc in range(col,col+3): ws.cell(row=3,column=cc).border = _border()
            col += 3
    ws.row_dimensions[3].height = 18

    # Fila 4: diámetros
    ws.cell(row=4,column=1,value="DIÁMETRO NOMINAL").font = _font(bold=True,color="1F2937")
    ws.cell(row=4,column=1).fill = _fill("D6E4F5")
    ws.cell(row=4,column=1).alignment = _align(); ws.cell(row=4,column=1).border = _border()
    col = 2
    for _ in caudales:
        for mat in materiales:
            for d in diams_nom:
                c = ws.cell(row=4,column=col,value=d)
                c.font = _font(bold=True,color=mat_font[mat])
                c.fill = _fill(mat_fill[mat]); c.alignment = _align(); c.border = _border()
                col += 1
    ws.row_dimensions[4].height = 16

    # Filas 5-7: calculadas
    for ri, (label, key) in enumerate([
        ("VELOCIDAD POR CONTINUIDAD (m/s)","Velocidad"),
        ("PÉRDIDA DE CARGA H-W (m)","Hf_HW"),
        ("PÉRDIDA DE CARGA D-W (m)","Hf_DW"),
    ], start=5):
        bg = XL["row_even"] if ri%2==0 else XL["row_odd"]
        c = ws.cell(row=ri,column=1,value=label)
        c.font = _font(bold=True,color="1F2937"); c.fill = _fill(bg)
        c.alignment = _align(h="left"); c.border = _border()
        col = 2
        for q in caudales:
            for mat in materiales:
                for d in diams_nom:
                    val = ""
                    if (q,mat,d) in idx:
                        v = idx[(q,mat,d)].get(key,"")
                        val = round(v,3) if isinstance(v,float) else v
                    c2 = ws.cell(row=ri,column=col,value=val)
                    c2.fill = _fill(bg); c2.border = _border()
                    c2.font = _font(color="1F2937"); c2.alignment = _align()
                    col += 1
        ws.row_dimensions[ri].height = 16

    # Filas 8-9: editadas (CFD y Experimental)
    for ri, (label, fk) in enumerate([
        ("PÉRDIDA DE CARGA OPEN FOAM (CFD)","cfd"),
        ("PÉRDIDA EN EXPERIMENTACIÓN","experimental"),
    ], start=8):
        bg = XL["row_even"] if ri%2==0 else XL["row_odd"]
        c = ws.cell(row=ri,column=1,value=label)
        c.font = _font(bold=True,color="1F2937"); c.fill = _fill(bg)
        c.alignment = _align(h="left"); c.border = _border()
        valores = datos_editados.get(fk, [""] * 27)
        for ci_idx in range(27):
            val = valores[ci_idx] if ci_idx < len(valores) else ""
            cell = ws.cell(row=ri, column=2+ci_idx, value=val if val else "")
            cell.fill = _fill(bg); cell.border = _border()
            cell.font = _font(color="1F2937"); cell.alignment = _align()
        ws.row_dimensions[ri].height = 16

    # Fila 10: espacio
    ws.row_dimensions[10].height = 6

    # Fila 11: planteamiento (fila completa)
    ws.merge_cells(f"A11:{last_col}11")
    c = ws.cell(row=11,column=1,
                value="Planteamiento del diseño experimental  (solo para caso longitudinal)")
    c.font = _font(bold=True,color="12294A")
    c.fill = _fill("D6E4F5"); c.alignment = _align(); c.border = _border()
    ws.row_dimensions[11].height = 18

    # Filas 12-14: réplicas con datos editados
    for ri2, (lbl, fk) in enumerate([
        ("PRIMERA RÉPLICA","replica1"),
        ("SEGUNDA RÉPLICA","replica2"),
        ("TERCERA RÉPLICA","replica3"),
    ], start=12):
        c = ws.cell(row=ri2,column=1,value=lbl)
        c.font = _font(bold=True,color="1F2937"); c.fill = _fill(XL["row_odd"])
        c.alignment = _align(h="left"); c.border = _border()
        valores = datos_editados.get(fk, [""] * 27)
        for ci_idx in range(27):
            val = valores[ci_idx] if ci_idx < len(valores) else ""
            cell = ws.cell(row=ri2, column=2+ci_idx, value=val if val else "")
            cell.fill = _fill(XL["row_odd"]); cell.border = _border()
            cell.font = _font(color="1F2937"); cell.alignment = _align()
        ws.row_dimensions[ri2].height = 16

    ws.column_dimensions["A"].width = 34
    for col in range(2, total_cols+1):
        ws.column_dimensions[get_column_letter(col)].width = 8


def build_excel(lista: list, datos_editados: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for i, r in enumerate(lista, 1):
        ws = wb.create_sheet(title=safe_sheet_name(f"Rep{i} {r['material']} {r['diam_nom']} {r['caudal_ls']}Ls"))
        build_detalle_sheet(ws, r, i)
    ws_res = wb.create_sheet(title="RESUMEN CONSOLIDADO")
    build_resumen_sheet(ws_res, lista, datos_editados)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PDF — recibe datos_editados
# ══════════════════════════════════════════════════════════════════════════════
def build_pdf(lista: list, datos_editados: dict) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, PageBreak,
                                        HRFlowable, NextPageTemplate,
                                        PageTemplate, Frame)
        from reportlab.lib.colors import HexColor
    except ImportError:
        return b""

    buf = io.BytesIO()
    C_DARK  = HexColor("#12294A"); C_MED  = HexColor("#1A4A7A")
    C_LIGHT = HexColor("#2563A8"); C_BG   = HexColor("#EDF3FA")
    C_PVC   = HexColor("#D6E4F5"); C_FG   = HexColor("#F5E6D0")
    C_HDPE  = HexColor("#D5EAD8"); C_WHITE = colors.white
    C_GREY  = HexColor("#4A6A8A")

    sty_title = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=15,
                                textColor=C_DARK, alignment=TA_CENTER, spaceAfter=4)
    sty_sub   = ParagraphStyle("s", fontName="Helvetica", fontSize=8.5,
                                textColor=C_GREY, alignment=TA_CENTER, spaceAfter=10)
    sty_h2    = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=10,
                                textColor=C_MED, spaceAfter=5, spaceBefore=12)

    pw, ph = A4
    lw, lh = landscape(A4)

    def _hf(canvas, doc, w, h):
        canvas.saveState()
        canvas.setFillColor(C_DARK); canvas.rect(0,h-26,w,26,fill=1,stroke=0)
        canvas.setFillColor(C_WHITE); canvas.setFont("Helvetica-Bold",8.5)
        canvas.drawString(1.5*cm, h-17, "ANÁLISIS DE PÉRDIDA DE CARGA — DISEÑO HIDRÁULICO")
        canvas.setFont("Helvetica",7.5)
        canvas.drawRightString(w-1.5*cm, h-17, f"Página {doc.page}")
        canvas.setStrokeColor(C_LIGHT); canvas.setLineWidth(1.2)
        canvas.line(0,h-27,w,h-27)
        canvas.setFillColor(C_GREY); canvas.setFont("Helvetica",6.5)
        canvas.drawCentredString(w/2,13,"Universidad Nacional del Altiplano · Puno 2026")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=1.8*cm)
    doc.addPageTemplates([
        PageTemplate(id='Portrait',  frames=[Frame(2*cm,1.8*cm,pw-4*cm,ph-2*cm-1.8*cm)],
                     onPage=lambda c,d:_hf(c,d,pw,ph), pagesize=A4),
        PageTemplate(id='Landscape', frames=[Frame(2*cm,1.8*cm,lw-4*cm,lh-2*cm-1.8*cm)],
                     onPage=lambda c,d:_hf(c,d,lw,lh), pagesize=landscape(A4)),
    ])

    story = [Spacer(1,1.2*cm)]
    story.append(Paragraph("ANÁLISIS DE PÉRDIDA DE CARGA", sty_title))
    story.append(Paragraph("Diseño Hidráulico — Modelamiento computacional de flujo interno en tuberías a presión", sty_sub))
    story.append(HRFlowable(width="100%",thickness=1.2,color=C_LIGHT,spaceAfter=5))
    story.append(Paragraph(f"Informe con {len(lista)} reporte(s) de cálculo", sty_sub))
    story.append(Spacer(1,0.5*cm))

    caudales  = [0.5,0.9,1.5]; materiales = ["PVC","FG","HDPE"]; diams_nom = ['1/2"','3/4"','1"']
    mat_col   = {"PVC":C_PVC,"FG":C_FG,"HDPE":C_HDPE}

    GS = lambda c: [("BACKGROUND",(0,0),(-1,0),C_MED),("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
                    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.5),
                    ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(0,1),(0,-1),"LEFT"),
                    ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("GRID",(0,0),(-1,-1),0.3,HexColor("#8EAADB")),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_WHITE,C_BG]),("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
                    ("ROWHEIGHT",(0,0),(-1,-1),15),("LEFTPADDING",(0,0),(-1,-1),7),
                    ("BACKGROUND",(0,1),(-1,-1),c)]

    for r in lista:
        story.append(PageBreak())
        story.append(Paragraph("REPORTE DETALLADO DE CÁLCULO HIDRÁULICO", sty_title))
        story.append(Paragraph(
            f"Tubería: <b>{r['nombre_tubo']}</b> · Material: <b>{r['material']}</b> · "
            f"Ø: <b>{r['diam_nom']}</b> · Q: <b>{r['caudal_ls']} L/s</b> · "
            f"L: <b>{r['longitud']:.2f} m</b> · T: <b>{r['temp']}°C</b>", sty_sub))
        story.append(HRFlowable(width="100%",thickness=0.8,color=C_LIGHT,spaceAfter=8))
        story.append(Paragraph("Parámetros Hidráulicos", sty_h2))
        det = [["Parámetro","Valor","Unidad","Referencia"]]
        for row in [
            ("Caudal",f"{r['caudal_ls']} L/s = {r['caudal_ls']/1000:.4f} m³/s","L/s / m³/s",""),
            ("Diámetro interno",f"{r['diametro_m']*1000:.1f} mm = {r['diametro_m']:.4f} m","mm / m",""),
            ("Área transversal",f"{r['area_m2']:.5f}","m²",""),
            ("Longitud",f"{r['longitud']:.2f}","m",""),
            ("Temperatura",f"{r['temp']}","°C",""),
            ("Viscosidad cinemática",f"{r['visc_cin']:.8f}","m²/s","Tablas estándar"),
            ("Velocidad media",f"{r['Velocidad']:.4f}","m/s","v = Q/A"),
            ("Número de Reynolds",f"{r['Re']:.0f}","—","Re = v·D/ν"),
            ("Rugosidad ε",f"{r['rugosidad']:.7f}","m",""),
            ("Relación ε/D",f"{r['rel_eD']:.7f}","—",""),
            ("Factor fricción f",f"{r['f']:.5f}","—","Swamee-Jain"),
            ("C Hazen-Williams",f"{r['C_H']:.3f}","—","Ec. regresión"),
            ("K Hazen-Williams",f"{r['K_HW']:.4f}","s^1.852/m2",""),
            ("Hf Hazen-Williams",f"{r['Hf_HW']:.5f}","m","Hf=K*Q^1.852"),
            ("K Darcy-Weisbach",f"{r['K_DW']:.4f}","s2/m5",""),
            ("Hf Darcy-Weisbach",f"{r['Hf_DW']:.5f}","m","Hf=K·Q²"),
        ]: det.append(list(row))
        t = Table(det, colWidths=[5.5*cm,4.5*cm,3*cm,3.5*cm], repeatRows=1)
        mc = mat_col.get(r.get("material",""), C_BG)
        t.setStyle(TableStyle(GS(C_WHITE)))
        story.append(t)
        story.append(Spacer(1,0.4*cm))
        story.append(Paragraph("Comparación de Métodos", sty_h2))
        tc = Table([
            ["Método","Q (m³/s)","v (m/s)","Coeficiente","K","Hf (m)"],
            ["Hazen-Williams",f"{r['caudal_ls']/1000:.4f}",f"{r['Velocidad']:.3f}",
             f"{r['C_H']:.3f}",f"{r['K_HW']:.4f}",f"{r['Hf_HW']:.5f}"],
            ["Darcy-Weisbach",f"{r['caudal_ls']/1000:.4f}",f"{r['Velocidad']:.3f}",
             f"{r['f']:.4f}",f"{r['K_DW']:.4f}",f"{r['Hf_DW']:.5f}"],
        ], colWidths=[4.5*cm,2.5*cm,2.5*cm,3*cm,2.5*cm,3*cm], repeatRows=1)
        tc.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),C_MED),("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("GRID",(0,0),(-1,-1),0.3,HexColor("#8EAADB")),
            ("BACKGROUND",(0,1),(-1,-1),mc),("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
            ("ROWHEIGHT",(0,0),(-1,-1),17),
        ]))
        story.append(tc)

    # Resumen landscape con datos editados
    story.append(NextPageTemplate('Landscape'))
    story.append(PageBreak())
    story.append(Paragraph("CUADRO CONSOLIDADO DE RESULTADOS", sty_title))
    story.append(HRFlowable(width="100%",thickness=1.2,color=C_LIGHT,spaceAfter=8))

    idx = {(r["caudal_ls"],r["material"],r["diam_nom"]):r for r in lista}
    mat_c = {"PVC":C_PVC,"FG":C_FG,"HDPE":C_HDPE}
    mat_t = {"PVC":HexColor("#0D2137"),"FG":HexColor("#5A2D00"),"HDPE":HexColor("#1A3D1F")}

    def calc_row(label, key):
        row = [label]
        for q in caudales:
            for mat in materiales:
                for d in diams_nom:
                    val = ""
                    if key and (q,mat,d) in idx:
                        v = idx[(q,mat,d)].get(key,"")
                        val = f"{v:.3f}" if isinstance(v,float) else str(v)
                    row.append(val)
        return row

    def edit_row(label, fk):
        row = [label]
        vals = datos_editados.get(fk, [""] * 27)
        for v in vals[:27]:
            row.append(v if v else "")
        return row

    all_data = [
        ["CAUDAL (L/S)"] + [str(q) for q in caudales for _ in range(9)],
        ["MATERIAL"]     + [m for _ in caudales for m in materiales for _ in range(3)],
        ["Ø NOMINAL"]    + [d for _ in caudales for _ in materiales for d in diams_nom],
        calc_row("VELOCIDAD (m/s)","Velocidad"),
        calc_row("Hf H-W (m)","Hf_HW"),
        calc_row("Hf D-W (m)","Hf_DW"),
        edit_row("FOAM/CFD","cfd"),
        edit_row("EXPERIMENTAL","experimental"),
        ["Planteamiento experimental (caso longitudinal)"]+[""]*27,
        edit_row("PRIMERA RÉPLICA","replica1"),
        edit_row("SEGUNDA RÉPLICA","replica2"),
        edit_row("TERCERA RÉPLICA","replica3"),
    ]

    cw = [4.2*cm]+[0.78*cm]*27
    t2 = Table(all_data, colWidths=cw, repeatRows=3)
    bs = [
        ("FONTSIZE",(0,0),(-1,-1),6.0),("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("GRID",(0,0),(-1,-1),0.3,HexColor("#8EAADB")),
        ("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("BACKGROUND",(0,0),(0,0),C_DARK),("TEXTCOLOR",(0,0),(0,0),C_WHITE),
        ("FONTNAME",(0,0),(0,0),"Helvetica-Bold"),
        ("BACKGROUND",(1,0),(-1,0),C_MED),("TEXTCOLOR",(1,0),(-1,0),C_WHITE),
        ("FONTNAME",(1,0),(-1,0),"Helvetica-Bold"),
        ("BACKGROUND",(0,1),(0,2),HexColor("#D6E4F5")),
        ("FONTNAME",(0,1),(0,2),"Helvetica-Bold"),("TEXTCOLOR",(0,1),(0,2),C_DARK),
        ("FONTNAME",(0,3),(0,11),"Helvetica-Bold"),("ALIGN",(0,3),(0,11),"LEFT"),
        ("BACKGROUND",(0,3),(0,7),HexColor("#D6E4F5")),("TEXTCOLOR",(0,3),(0,7),C_DARK),
        ("BACKGROUND",(0,8),(-1,8),HexColor("#D6E4F5")),("TEXTCOLOR",(0,8),(-1,8),C_DARK),
        ("FONTNAME",(0,8),(-1,8),"Helvetica-Bold"),
        ("BACKGROUND",(0,9),(-1,11),HexColor("#EDF3FA")),("TEXTCOLOR",(0,9),(-1,11),C_DARK),
        ("ROWBACKGROUNDS",(1,3),(-1,7),[C_WHITE,C_BG]),("TEXTCOLOR",(1,3),(-1,7),C_DARK),
        ("TEXTCOLOR",(1,9),(-1,11),C_DARK),
    ]
    col = 1
    for _ in caudales:
        for mat in materiales:
            for ri in [1,2]:
                for cc in range(col,col+3):
                    bs += [("BACKGROUND",(cc,ri),(cc,ri),mat_c[mat]),
                           ("TEXTCOLOR",(cc,ri),(cc,ri),mat_t[mat]),
                           ("FONTNAME",(cc,ri),(cc,ri),"Helvetica-Bold")]
            col += 3
    t2.setStyle(TableStyle(bs))
    story.append(t2)
    doc.build(story)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PESTAÑAS
# ══════════════════════════════════════════════════════════════════════════════
tabs = st.tabs(["  DATOS INICIALES","  DATOS DE ENTRADA","  RESULTADOS","  HOJA DE RESUMEN"])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 0
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[0]:
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)
    col_img, col_ctrl = st.columns([1, 2], gap="large")

    with col_ctrl:
        st.markdown('<div class="dh-card">', unsafe_allow_html=True)
        st.markdown('<div class="dh-card-title">Parámetros de control</div>', unsafe_allow_html=True)
        st.session_state.longitud = st.slider("Longitud del tramo (L) — metros", 1.0, 10.0, 1.5, step=0.1)
        st.session_state.temp = st.select_slider("Temperatura del agua (T) — °C",
            options=[0,5,10,15,20,25,30,40,50,60,70,80,90,100], value=10)
        fila_agua = df_agua.loc[df_agua['Temp'] == st.session_state.temp].iloc[0]
        st.session_state.visc_cin = float(fila_agua['Visc_Cinem'])
        st.markdown("<hr class='dh-sep'>", unsafe_allow_html=True)
        m1, m2, m3 = st.columns(3)
        with m1:
            st.markdown(f"""<div class="dh-metric"><div class="dh-metric-label">Longitud</div>
                <div class="dh-metric-value">{st.session_state.longitud:.2f}</div>
                <div class="dh-metric-unit">metros</div></div>""", unsafe_allow_html=True)
        with m2:
            st.markdown(f"""<div class="dh-metric"><div class="dh-metric-label">Temperatura</div>
                <div class="dh-metric-value">{st.session_state.temp}</div>
                <div class="dh-metric-unit">°C</div></div>""", unsafe_allow_html=True)
        with m3:
            st.markdown(f"""<div class="dh-metric"><div class="dh-metric-label">Visc. Cinemática</div>
                <div class="dh-metric-value">{st.session_state.visc_cin:.2e}</div>
                <div class="dh-metric-unit">m²/s</div></div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_img:
        st.markdown('<div class="dh-card">', unsafe_allow_html=True)
        st.markdown('<div class="dh-card-title">Esquema del sistema</div>', unsafe_allow_html=True)
        if os.path.exists("Imagen2.png"):
            st.image("Imagen2.png", use_container_width=True)
        else:
            st.info("Coloca 'Imagen2.png' en el directorio del proyecto.")
        st.markdown('</div>', unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 1
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[1]:
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)

    if st.session_state.get("go_to_tab") == "banner":
        st.success("✅ Resultado guardado. Ingresa nuevos parámetros y presiona CALCULAR.")
        st.session_state.go_to_tab = None

    st.markdown('<div class="dh-card">', unsafe_allow_html=True)
    st.markdown('<div class="dh-card-title">Configuración hidráulica</div>', unsafe_allow_html=True)

    col_left, col_right = st.columns(2, gap="large")
    with col_left:
        nombre_tubo = st.selectbox("Seleccione tubería:", df_tubo['Nombre'].unique())
        fila_tubo   = df_tubo[df_tubo['Nombre'] == nombre_tubo].iloc[0]
        material    = detectar_material_tubo(nombre_tubo, df_rug)
        filas_mat_r = df_rug[df_rug['Material'] == material]
        rugosidad   = float(filas_mat_r.iloc[0]['Rugosidad']) if not filas_mat_r.empty else 0.0000015
        d_m  = float(fila_tubo['Diametro_m'])
        a_m2 = float(fila_tubo['Area_m2'])
        st.markdown(f"""<div class="dh-metric" style="margin-top:12px">
            <div class="dh-metric-label">Material detectado · Rugosidad ε</div>
            <div class="dh-metric-value" style="font-size:1rem">{material}</div>
            <div class="dh-metric-unit">{rugosidad:.7f} </div>
        </div>""", unsafe_allow_html=True)

    with col_right:
        caudal_ls = st.radio("Caudal de diseño (L/s):", [0.5, 0.9, 1.5], horizontal=True)
        st.markdown(f"""<div class="dh-metric" style="margin-top:12px">
            <div class="dh-metric-label">Diámetro interno · Área transversal</div>
            <div class="dh-metric-value" style="font-size:1rem">{d_m*1000:.2f} mm</div>
            <div class="dh-metric-unit">{a_m2:.6f} m²</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

    btn_col1, btn_col2 = st.columns(2, gap="medium")
    with btn_col1:
        if st.button("⚡  CALCULAR RESULTADOS", use_container_width=True, type="primary"):
            st.session_state.update({
                "nombre_tubo": nombre_tubo, "material": material,
                "rugosidad": rugosidad, "diametro": d_m, "area": a_m2,
                "caudal_ls": caudal_ls, "visc_cin": st.session_state.get("visc_cin", 1e-6),
            })
            st.success("✅ Cálculos listos — navega a la pestaña RESULTADOS.")

    with btn_col2:
        if st.button("🗂  PROCESAR TODAS LAS COMBINACIONES", use_container_width=True):
            longitud = st.session_state.get("longitud", 1.5)
            temp_val = st.session_state.get("temp", 10)
            visc_val = st.session_state.get("visc_cin", 1e-6)
            count = 0
            for q in [0.5, 0.9, 1.5]:
                for mat in ["PVC", "FG", "HDPE"]:
                    filas_r = df_rug[df_rug['Material'] == mat]
                    if filas_r.empty: continue
                    rug_mat  = float(filas_r.iloc[0]['Rugosidad'])
                    # FIX: usar detectar_material_tubo para filtrar correctamente FG
                    tubos_mat = df_tubo[df_tubo['Nombre'].apply(
                        lambda n: detectar_material_tubo(n, df_rug) == mat)]
                    if tubos_mat.empty: continue
                    for _, row_t in tubos_mat.iterrows():
                        r = calcular(
                            nombre_tubo=row_t['Nombre'], material=mat, rugosidad=rug_mat,
                            diametro=float(row_t['Diametro_m']), area=float(row_t['Area_m2']),
                            caudal_ls=q, longitud=longitud, temp=temp_val, visc_cin=visc_val
                        )
                        key = (r["caudal_ls"], r["material"], r["diam_nom"])
                        existing = [(x["caudal_ls"],x["material"],x["diam_nom"])
                                    for x in st.session_state.lista_resultados]
                        if key not in existing:
                            st.session_state.lista_resultados.append(r)
                            count += 1
            if count > 0:
                st.success(f"✅ {count} combinaciones procesadas. Ve a HOJA DE RESUMEN.")
            else:
                st.info("Todas las combinaciones ya estaban procesadas.")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 2
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[2]:
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)

    if 'caudal_ls' not in st.session_state:
        st.markdown("""<div class="dh-card" style="text-align:center;padding:38px">
            <div style="font-family:'Rajdhani',sans-serif;font-size:1.3rem;color:#5a7aaa;letter-spacing:3px">
            ⚠ DATOS NO CALCULADOS</div>
            <div style="color:#7a8aaa;margin-top:8px;font-size:0.83rem">
            Ve a DATOS DE ENTRADA y presiona CALCULAR RESULTADOS</div>
        </div>""", unsafe_allow_html=True)
    else:
        Q_Ls=st.session_state.caudal_ls; Q_m3s=Q_Ls/1000
        D_m=st.session_state.diametro;   D_mm=D_m*1000
        A=st.session_state.area;         eps=st.session_state.rugosidad
        L=st.session_state.longitud;     v_cin=st.session_state.visc_cin
        mat=st.session_state.get("material","PVC")
        V=Q_m3s/A; Re=(D_m*V)/v_cin; eD=eps/D_m
        f=0.25/(np.log10((eps/(3.71*D_m))+(5.74/(Re**0.9))))**2
        lnRe=np.log(Re); lneD=np.log(eD)
        C_H=(197.17-25.79*lnRe-5.41*lneD+0.4464*lnRe**2
             -3.39*lneD**2-5.086*lnRe*lneD+0.041*lnRe**3
             +0.124*lneD**3+0.39*lnRe*lneD**2+0.3757*lnRe**2*lneD)
        K_HW=(10.67*L)/((C_H**1.852)*(D_m**4.87)); Hf_HW=K_HW*(Q_m3s**1.852)
        K_DW=(0.08263*f*L)/(D_m**5);               Hf_DW=K_DW*(Q_m3s**2)
        diam_nom=detectar_diam_nom(st.session_state.nombre_tubo)

        mc_hex={"PVC":"#1a56b0","FG":"#8A5200","HDPE":"#1A7A40"}.get(mat,"#2563A8")
        st.markdown(f"""<div style="display:flex;align-items:center;gap:14px;margin-bottom:18px">
            <div style="background:{mc_hex};padding:7px 18px;font-family:'Rajdhani',sans-serif;
            font-size:0.95rem;font-weight:700;letter-spacing:2px;color:white;border-radius:2px;">
            {st.session_state.get('nombre_tubo','—')}</div>
            <div style="font-family:'Share Tech Mono',monospace;color:#5a7aaa;font-size:0.72rem">
            Q={Q_Ls} L/s · L={L:.2f} m · T={st.session_state.get('temp',10)}°C</div>
        </div>""", unsafe_allow_html=True)

        mc1,mc2,mc3,mc4=st.columns(4)
        for cm,(label,unit,val) in zip([mc1,mc2,mc3,mc4],[
            ("Velocidad","m/s",f"{V:.4f}"),("Reynolds","—",f"{Re:,.0f}"),
            ("Hf Hazen-W.","m",f"{Hf_HW:.5f}"),("Hf Darcy-W.","m",f"{Hf_DW:.5f}")]):
            with cm:
                st.markdown(f"""<div class="dh-metric">
                    <div class="dh-metric-label">{label}</div>
                    <div class="dh-metric-value">{val}</div>
                    <div class="dh-metric-unit">{unit}</div></div>""", unsafe_allow_html=True)

        st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)
        col_t, col_i = st.columns([3,2], gap="large")

        with col_t:
            st.markdown('<div class="dh-card">', unsafe_allow_html=True)
            st.markdown('<div class="dh-card-title">Parámetros detallados</div>', unsafe_allow_html=True)
            filas=[
                ("Caudal",f"<b style='color:#c8deff'>{Q_Ls}</b> L/s = {Q_m3s:.4f} m³/s"),
                ("Diámetro interior",f"<b style='color:#c8deff'>{D_mm:.1f}</b> mm = {D_m:.4f} m"),
                ("Área transversal",f"{A:.5f} m²"),("Longitud",f"{L:.2f} m"),
                ("Velocidad media",f"{V:.4f} m/s"),
                ("Temperatura",f"{st.session_state.get('temp',10)} °C"),
                ("Viscosidad cinemática",f"{v_cin:.8f} m²/s"),
                ("Número de Reynolds",f"{Re:,.0f}"),
                ("Rugosidad ε",f"{eps:.7f} m"),("Relación ε/D",f"{eD:.7f}"),
                ("Factor fricción f",f"{f:.5f}"),("C Hazen-Williams",f"{C_H:.3f}"),
            ]
            html_t='<table class="res-table"><thead><tr><th>PARÁMETRO</th><th>VALOR</th></tr></thead><tbody>'
            for label,val in filas:
                html_t+=f'<tr><td class="left-align">{label}</td><td>{val}</td></tr>'
            html_t+='</tbody></table>'
            st.markdown(html_t, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        with col_i:
            st.markdown('<div class="dh-card">', unsafe_allow_html=True)
            st.markdown('<div class="dh-card-title">Diagrama de flujo</div>', unsafe_allow_html=True)
            if os.path.exists("Imagen3.png"):
                st.image("Imagen3.png", use_container_width=True)
            else:
                st.info("Coloca 'Imagen3.png' en el directorio.")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="dh-card">', unsafe_allow_html=True)
        st.markdown('<div class="dh-card-title">Comparación de métodos</div>', unsafe_allow_html=True)
        html_comp="""<table class="res-table"><thead><tr>
            <th>MÉTODO</th><th>Q (m³/s)</th><th>v (m/s)</th>
            <th>LONGITUD (m)</th><th>COEFICIENTE</th><th>K</th><th>Hf (m)</th>
        </tr></thead><tbody>"""
        html_comp+=f"""<tr><td class="left-align">Hazen-Williams</td>
            <td>{Q_m3s:.4f}</td><td>{V:.3f}</td><td>{L:.2f}</td>
            <td>C = {C_H:.3f}</td><td>{K_HW:.4f}</td>
            <td class="highlight-hw">{Hf_HW:.5f}</td></tr>"""
        html_comp+=f"""<tr><td class="left-align">Darcy-Weisbach</td>
            <td>{Q_m3s:.4f}</td><td>{V:.3f}</td><td>{L:.2f}</td>
            <td>f = {f:.5f}</td><td>{K_DW:.4f}</td>
            <td class="highlight-dw">{Hf_DW:.5f}</td></tr>"""
        html_comp+="</tbody></table>"
        st.markdown(html_comp, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<hr class="dh-sep">', unsafe_allow_html=True)
        ba1,ba2,ba3=st.columns(3, gap="medium")
        with ba1:
            if st.button("💾  Guardar en Hoja de Resumen", type="primary", use_container_width=True):
                st.session_state.lista_resultados.append({
                    "nombre_tubo":st.session_state.nombre_tubo,"material":mat,
                    "diam_nom":diam_nom,"caudal_ls":Q_Ls,"diametro_m":D_m,
                    "area_m2":A,"longitud":L,"temp":st.session_state.get("temp",10),
                    "visc_cin":v_cin,"rugosidad":eps,"Velocidad":V,"Re":Re,
                    "rel_eD":eD,"f":f,"C_H":C_H,"K_HW":K_HW,"Hf_HW":Hf_HW,
                    "K_DW":K_DW,"Hf_DW":Hf_DW,
                })
                st.success(f"✅ Reporte {len(st.session_state.lista_resultados)} guardado.")
        with ba2:
            if st.button("➕  Añadir otra iteración", use_container_width=True):
                for key in ['caudal_ls','diametro','longitud','temp',
                            'rugosidad','visc_cin','nombre_tubo','area','material']:
                    st.session_state.pop(key, None)
                st.session_state.go_to_tab=1; st.rerun()
        with ba3:
            if st.button("🖨  Imprimir página", use_container_width=True):
                st.components.v1.html("<script>window.parent.window.print();</script>", height=0)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB 3
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tabs[3]:
    st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)
    lista=st.session_state.lista_resultados
    caudales=[0.5,0.9,1.5]; materiales=["PVC","FG","HDPE"]; diams_nom=['1/2"','3/4"','1"']
    idx={(r["caudal_ls"],r["material"],r["diam_nom"]):r for r in lista}
    cm_map={"PVC":"#bdd6ee","FG":"#f5e0c0","HDPE":"#c6e8cc"}
    cf_map={"PVC":"#0d2a50","FG":"#5a2d00","HDPE":"#1a3d1f"}
    n_cols=27

    for fk in ["cfd","experimental","replica1","replica2","replica3"]:
        if fk not in st.session_state:
            st.session_state[fk]=[""] * n_cols
    if "edit_mode_resumen" not in st.session_state:
        st.session_state.edit_mode_resumen=False

    def th(text, bg="#1A4A7A", color="#FFFFFF", cs=1):
        return (f'<th colspan="{cs}" style="background:{bg};color:{color};'
                f'padding:7px 4px;text-align:center;border:1px solid #c5d5ea;'
                f'font-family:Rajdhani,sans-serif;font-size:10px;letter-spacing:1px;">{text}</th>')
    def td2(text, bg="#ffffff", bold=False, align="center", color="#1a2a4a"):
        fw="font-weight:700;" if bold else ""
        return (f'<td style="background:{bg};{fw}color:{color};'
                f'padding:6px 4px;text-align:{align};border:1px solid #c5d5ea;font-size:9.5px;">{text}</td>')

    html=('<div style="overflow-x:auto;border:1px solid #2a3550;border-radius:4px;">'
          '<table style="border-collapse:collapse;width:100%;font-family:Exo 2,sans-serif;">')
    html+=f'<tr>{th("CAUDAL (L/S)",bg="#1A4A7A",color="#FFFFFF")}'
    for q in caudales: html+=th(str(q),bg="#2563A8",color="#FFFFFF",cs=9)
    html+='</tr>'
    html+=f'<tr>{th("MATERIAL",bg="#1A4A7A",color="#FFFFFF")}'
    for _ in caudales:
        for mat in materiales: html+=th(mat,bg=cm_map[mat],color=cf_map[mat],cs=3)
    html+='</tr>'
    html+=f'<tr>{th("DIÁMETRO NOMINAL",bg="#1A4A7A",color="#FFFFFF")}'
    for _ in caudales:
        for mat in materiales:
            for d in diams_nom: html+=th(d,bg=cm_map[mat],color=cf_map[mat])
    html+='</tr>'

    for ri,(label,key) in enumerate([
        ("VELOCIDAD POR CONTINUIDAD (m/s)","Velocidad"),
        ("PÉRDIDA DE CARGA H-W (m)","Hf_HW"),
        ("PÉRDIDA DE CARGA D-W (m)","Hf_DW"),
    ]):
        rbg=["#f0f5fb","#ffffff"][ri%2]
        html+=f'<tr>{td2(label,bg="#eef4fb",bold=True,align="left",color="#1a3a6a")}'
        for q in caudales:
            for mat in materiales:
                for d in diams_nom:
                    val=""
                    if (q,mat,d) in idx:
                        v=idx[(q,mat,d)].get(key,"")
                        val=f"{v:.3f}" if isinstance(v,float) else str(v)
                    html+=td2(val,bg=rbg,color="#1a56b0" if val else "#aabbcc")
        html+='</tr>'

    for label,fk in [("PÉRDIDA DE CARGA OPEN FOAM (CFD)","cfd"),
                      ("PÉRDIDA EN EXPERIMENTACIÓN","experimental")]:
        html+=f'<tr>{td2(label,bg="#eef4fb",bold=True,align="left",color="#1a3a6a")}'
        for ci in range(n_cols):
            val=st.session_state[fk][ci] if ci<len(st.session_state[fk]) else ""
            html+=td2(val if val else "—",bg="#f5f8fd",color="#3a6a8a")
        html+='</tr>'

    # Planteamiento — colspan=28 (columna etiqueta + 27 datos)
    html+=(f'<tr><td colspan="28" style="background:#ddeaf8;color:#1a3a6a;'
           f'padding:8px 12px;text-align:left;border:1px solid #c5d5ea;'
           f'font-family:Rajdhani,sans-serif;font-size:10px;font-weight:700;letter-spacing:1px;">'
           f'Planteamiento del diseño experimental &nbsp;(solo para caso longitudinal)'
           f'</td></tr>')

    for label,fk in [("PRIMERA RÉPLICA","replica1"),("SEGUNDA RÉPLICA","replica2"),("TERCERA RÉPLICA","replica3")]:
        html+=f'<tr>{td2(label,bg="#eef4fb",bold=True,align="left",color="#1a3a6a")}'
        for ci in range(n_cols):
            val=st.session_state[fk][ci] if ci<len(st.session_state[fk]) else ""
            html+=td2(val if val else "—",bg="#ffffff",color="#3a6a8a")
        html+='</tr>'
    html+='</table></div>'

    st.markdown('<div class="dh-card">', unsafe_allow_html=True)
    st.markdown('<div class="dh-card-title">Cuadro de resultados acumulados</div>', unsafe_allow_html=True)
    st.markdown(html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if len(lista):
        st.markdown(f"""<div style="text-align:center;font-family:'Share Tech Mono',monospace;
            color:#1a56b0;font-size:0.72rem;letter-spacing:2px;margin:6px 0">
            {len(lista)} REPORTE(S) REGISTRADO(S)</div>""", unsafe_allow_html=True)

    st.markdown('<hr class="dh-sep">', unsafe_allow_html=True)

    if st.button(
        "✏️  Editar datos manuales" if not st.session_state.edit_mode_resumen else "💾  Guardar cambios",
        use_container_width=False,
        type="primary" if st.session_state.edit_mode_resumen else "secondary",
        key="btn_edit_resumen"
    ):
        st.session_state.edit_mode_resumen=not st.session_state.edit_mode_resumen; st.rerun()

    if st.session_state.edit_mode_resumen:
        st.markdown("""<div style="background:rgba(26,86,176,0.06);border:1px solid #1a56b0;
            border-radius:4px;padding:12px 16px;margin:10px 0;
            font-family:'Share Tech Mono',monospace;font-size:0.7rem;color:#1a56b0;letter-spacing:1px;">
            ⚠ MODO EDICIÓN ACTIVO — Modifica las filas manuales y presiona <b>Guardar cambios</b>.
        </div>""", unsafe_allow_html=True)

        col_headers=[]
        for q in caudales:
            for mat in materiales:
                for d in diams_nom:
                    col_headers.append(f"Q={q}|{mat}|{d}")

        for fk,flabel in [("cfd","Open Foam (CFD)"),("experimental","Experimentación")]:
            st.markdown(f'<div class="dh-card" style="margin-top:10px">', unsafe_allow_html=True)
            st.markdown(f'<div class="dh-card-title">Pérdida de Carga — {flabel}</div>', unsafe_allow_html=True)
            cols_e=st.columns(9)
            for ci in range(n_cols):
                with cols_e[ci%9]:
                    nv=st.text_input(col_headers[ci],value=st.session_state[fk][ci],key=f"{fk}_{ci}")
                    st.session_state[fk][ci]=nv
            st.markdown('</div>', unsafe_allow_html=True)

        for rep_label,rep_key in [("PRIMERA RÉPLICA","replica1"),
                                   ("SEGUNDA RÉPLICA","replica2"),
                                   ("TERCERA RÉPLICA","replica3")]:
            st.markdown('<div class="dh-card" style="margin-top:8px">', unsafe_allow_html=True)
            st.markdown(f'<div class="dh-card-title">{rep_label}</div>', unsafe_allow_html=True)
            rc=st.columns(9)
            for ci in range(n_cols):
                with rc[ci%9]:
                    nv=st.text_input(col_headers[ci],value=st.session_state[rep_key][ci],key=f"{rep_key}_{ci}")
                    st.session_state[rep_key][ci]=nv
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Exportaciones — datos_editados se lee en el momento de exportar ───────
    st.markdown('<hr class="dh-sep">', unsafe_allow_html=True)
    datos_editados={
        "cfd":          list(st.session_state.get("cfd",          [""] * n_cols)),
        "experimental": list(st.session_state.get("experimental", [""] * n_cols)),
        "replica1":     list(st.session_state.get("replica1",     [""] * n_cols)),
        "replica2":     list(st.session_state.get("replica2",     [""] * n_cols)),
        "replica3":     list(st.session_state.get("replica3",     [""] * n_cols)),
    }

    if lista:
        eb1,eb2,eb3=st.columns(3, gap="medium")
        with eb1:
            xl_bytes=build_excel(lista, datos_editados)
            st.download_button("📊  Exportar Excel", data=xl_bytes,
                file_name="AnalisisPerdidaCarga.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with eb2:
            try:
                pdf_bytes=build_pdf(lista, datos_editados)
                if pdf_bytes:
                    st.download_button("📄  Exportar PDF", data=pdf_bytes,
                        file_name="AnalisisPerdidaCarga.pdf",
                        mime="application/pdf", use_container_width=True)
                else:
                    st.info("Instala reportlab: pip install reportlab")
            except Exception as e:
                st.error(f"Error PDF: {e}")
        with eb3:
            if st.button("🗑  Limpiar todo", use_container_width=True):
                st.session_state.lista_resultados=[]
                for fk in ["cfd","experimental","replica1","replica2","replica3"]:
                    st.session_state[fk]=[""] * n_cols
                st.rerun()
    else:
        st.markdown("""<div style="text-align:center;padding:28px;color:#7a8aaa;
            font-family:'Share Tech Mono',monospace;font-size:0.77rem;letter-spacing:2px">
            NINGÚN REPORTE GUARDADO AÚN<br>
            <span style="font-size:0.63rem">Ve a RESULTADOS → 💾 Guardar en Hoja de Resumen</span>
        </div>""", unsafe_allow_html=True)
